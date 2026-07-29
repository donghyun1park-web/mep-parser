"""
boq_export.py  —  geometry.json → 물량집계(BOQ) Excel

파싱된 기하에서 견적용 물량을 결정론적으로 집계한다:
  벽    : 두께별 — 개수 · 총길이(m) · 면적(㎡, 길이×높이) · 체적(㎥)
  기둥  : 단면별(원형 D촉/사각 W×D) — 개수 · 높이 · 콘크리트 체적
  슬래브: 개수 · 면적(㎡, 신발끈 공식) · 체적(㎥)
  창호  : door/window 규격별 개수
  MEP   : pipe/duct/tray 규격별 총길이(m)

Excel 양식은 schedule_io.export_schedule_xlsx 스타일(헤더 채움/테두리/열폭)을 따른다.

사용:
    python boq_export.py geometry.json            # <입력>_물량.xlsx
    python boq_export.py geometry.json out.xlsx
"""
import argparse
import json
import math
import os
import sys


def _poly_area(pts):
    """신발끈 공식 (mm² → 호출측에서 단위 변환)."""
    n = len(pts)
    if n < 3:
        return 0.0
    s = 0.0
    for i in range(n):
        x1, y1 = pts[i][0], pts[i][1]
        x2, y2 = pts[(i + 1) % n][0], pts[(i + 1) % n][1]
        s += x1 * y2 - x2 * y1
    return abs(s) / 2.0


def _polyline_len(pts):
    return sum(math.hypot(pts[i + 1][0] - pts[i][0], pts[i + 1][1] - pts[i][1])
               for i in range(len(pts) - 1))


def _bbox_wd(pts):
    xs = [p[0] for p in pts]
    ys = [p[1] for p in pts]
    return max(xs) - min(xs), max(ys) - min(ys)


def _r10(v):
    """10mm 반올림(제도 오차 흡수한 규격 그룹핑)."""
    return int(round(float(v) / 10.0) * 10)


def aggregate(data):
    """geometry.json dict → 섹션별 집계 rows. 반환 {섹션명: (헤더, rows, 합계행)}."""
    els = data.get("elements", {})
    params = data.get("params", {})
    pw = float(params.get("wall", {}).get("width", 200.0))
    ph = float(params.get("wall", {}).get("height", 2800.0))
    pcol_h = float(params.get("column", {}).get("height", 3000.0))
    pslab_t = float(params.get("slab", {}).get("thickness", 200.0))
    out = {}

    # ── 벽: 두께별 ───────────────────────────────────────────
    wg = {}
    for w in els.get("wall", []):
        ov = w.get("overrides") or {}
        h = float(ov.get("height", ph))
        if w.get("closed") and len(w.get("points", [])) >= 3:
            pts = w["points"]
            L = _polyline_len(pts + [pts[0]])
            vol = _poly_area(pts) * h            # 폐합벽: 단면적×높이
            key = "폐합(솔리드)"
            t = None
        else:
            cl = w.get("centerline") or w.get("points") or []
            if len(cl) < 2:
                continue
            L = _polyline_len(cl)
            t = float(w.get("width_detected") or ov.get("width") or pw)
            vol = L * h * t
            key = f"T{_r10(t)}"
        g = wg.setdefault(key, {"count": 0, "len": 0.0, "area": 0.0, "vol": 0.0,
                                "h": h})
        g["count"] += 1
        g["len"] += L
        g["area"] += L * h
        g["vol"] += vol
    rows = [[k, g["count"], round(g["len"] / 1000, 1), round(g["h"] / 1000, 2),
             round(g["area"] / 1e6, 1), round(g["vol"] / 1e9, 2)]
            for k, g in sorted(wg.items())]
    tot = ["합계", sum(r[1] for r in rows), round(sum(r[2] for r in rows), 1), "",
           round(sum(r[4] for r in rows), 1), round(sum(r[5] for r in rows), 2)]
    out["벽"] = (["두께", "개수", "길이(m)", "높이(m)", "면적(㎡)", "체적(㎥)"], rows, tot)

    # ── 기둥: 단면별 ─────────────────────────────────────────
    cg = {}
    for c in els.get("column", []):
        ov = c.get("overrides") or {}
        h = float(ov.get("height", pcol_h))
        if c.get("kind") == "circle":
            r = float(c.get("radius", 200.0))
            key = f"D{_r10(2 * r)}"
            area = math.pi * r * r
        elif c.get("points"):
            wd = _bbox_wd(c["points"])
            key = f"{_r10(wd[0])}x{_r10(wd[1])}"
            area = _poly_area(c["points"]) or wd[0] * wd[1]
        else:
            continue
        g = cg.setdefault(key, {"count": 0, "vol": 0.0, "h": h})
        g["count"] += 1
        g["vol"] += area * h
    rows = [[k, g["count"], round(g["h"] / 1000, 2), round(g["vol"] / 1e9, 2)]
            for k, g in sorted(cg.items())]
    tot = ["합계", sum(r[1] for r in rows), "", round(sum(r[3] for r in rows), 2)]
    out["기둥"] = (["단면(mm)", "개수", "높이(m)", "체적(㎥)"], rows, tot)

    # ── 슬래브 ──────────────────────────────────────────────
    rows = []
    ta = tv = 0.0
    for i, s in enumerate(els.get("slab", [])):
        if not s.get("points") or len(s["points"]) < 3:
            continue
        ov = s.get("overrides") or {}
        t = float(ov.get("thickness", pslab_t))
        a = _poly_area(s["points"])
        rows.append([f"Slab_{i}", s.get("layer", ""), round(a / 1e6, 1),
                     _r10(t), round(a * t / 1e9, 2)])
        ta += a
        tv += a * t
    tot = ["합계", "", round(ta / 1e6, 1), "", round(tv / 1e9, 2)]
    out["슬래브"] = (["번호", "레이어", "면적(㎡)", "두께(mm)", "체적(㎥)"], rows, tot)

    # ── 창호: 규격별 ─────────────────────────────────────────
    og = {}
    for o in els.get("opening", []):
        sub = o.get("subtype") or "opening"
        w = o.get("width")
        h = o.get("height")
        size = (f"{_r10(w)}x{_r10(h)}" if w and h else
                f"D{_r10(2 * o['radius'])}" if o.get("radius") else "-")
        key = ({"door": "문", "window": "창"}.get(sub, "개구부"), size)
        og[key] = og.get(key, 0) + 1
    rows = [[k[0], k[1], n] for k, n in sorted(og.items())]
    tot = ["합계", "", sum(r[2] for r in rows)]
    out["창호"] = (["구분", "규격(mm)", "개수"], rows, tot)

    # ── MEP: 규격별 길이 ─────────────────────────────────────
    mg = {}
    for cat, label in (("pipe", "배관"), ("duct", "덕트"), ("tray", "트레이")):
        for m in els.get(cat, []):
            pts = m.get("centerline") or m.get("points") or []
            if len(pts) < 2:
                continue
            ov = m.get("overrides") or {}
            size = ov.get("width") or m.get("width_detected") or \
                m.get("diameter") or ""
            size = f"{_r10(size)}" if size else "-"
            key = (label, size)
            g = mg.setdefault(key, {"count": 0, "len": 0.0})
            g["count"] += 1
            g["len"] += _polyline_len(pts)
    rows = [[k[0], k[1], g["count"], round(g["len"] / 1000, 1)]
            for k, g in sorted(mg.items())]
    tot = ["합계", "", sum(r[2] for r in rows), round(sum(r[3] for r in rows), 1)]
    out["MEP"] = (["구분", "규격(mm)", "개수", "길이(m)"], rows, tot)
    return out


def export_boq_xlsx(data, path, title=None):
    """geometry.json dict → 물량집계 Excel. schedule_io 스타일 재사용. 반환 경로."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    src = os.path.basename(data.get("source", "") or "geometry.json")
    title = title or f"물량집계 — {src}"
    secs = aggregate(data)

    wb = Workbook()
    ws = wb.active
    ws.title = "물량집계"
    hdr_fill = PatternFill("solid", fgColor="2D6CDF")
    sec_fill = PatternFill("solid", fgColor="EDF2FB")
    hdr_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="B0B7C3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:F1")
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(size=14, bold=True)
    t.alignment = Alignment(horizontal="center")
    qa = data.get("qa") or {}
    meta = f"파싱 커버리지 {qa.get('face_coverage_pct', '—')}%  ·  단위: m/㎡/㎥"
    ws.merge_cells("A2:F2")
    m = ws.cell(row=2, column=1, value=meta)
    m.font = Font(italic=True, color="7A8290")
    m.alignment = Alignment(horizontal="center")

    r = 4
    for name, (headers, rows, tot) in secs.items():
        if not rows:
            continue
        sc = ws.cell(row=r, column=1, value=f"■ {name}")
        sc.font = Font(bold=True, size=12)
        sc.fill = sec_fill
        r += 1
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        r += 1
        for row in rows + [tot]:
            bold = row is tot
            for c, v in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="left" if c == 1 else "right")
                if bold:
                    cell.font = Font(bold=True)
            r += 1
        r += 1  # 섹션 사이 빈 행

    for col, w in zip("ABCDEF", [16, 10, 12, 10, 12, 12]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"
    wb.save(path)
    return path


def main():
    ap = argparse.ArgumentParser(description="geometry.json → 물량집계 Excel")
    ap.add_argument("geometry", help="geometry.json 경로")
    ap.add_argument("out", nargs="?", default=None,
                    help="출력 .xlsx (기본 <입력>_물량.xlsx)")
    args = ap.parse_args()
    with open(args.geometry, encoding="utf-8") as f:
        data = json.load(f)
    out = args.out or os.path.splitext(args.geometry)[0] + "_물량.xlsx"
    try:
        export_boq_xlsx(data, out)
    except ImportError:
        print("[ERROR] openpyxl 필요: pip install openpyxl", file=sys.stderr)
        sys.exit(1)
    secs = aggregate(data)
    parts = []
    for name, (_, rows, tot) in secs.items():
        if rows:
            parts.append(f"{name} {tot[1] if isinstance(tot[1], int) else len(rows)}")
    print(f"[OK] 물량집계 -> {out}  ({', '.join(parts)})")


if __name__ == "__main__":
    main()
