"""
schedule_io.py — 창호일람(window/door schedule) Excel 양식 입출력

목적:
  - 평면도(DXF)에는 창의 '높이·종류'가 없다(평면엔 폭·위치만). 그 정보는 창호일람표에만 있다.
  - 이 모듈은 창호일람을 **Excel 양식(.xlsx)** 으로 주고받게 한다.
      export_schedule_xlsx(): 파싱으로 추출한 일람(또는 빈 양식)을 Excel 로 저장.
      load_schedule_xlsx():   사람이 채운/수정한 Excel 을 다시 schedule 레코드로 읽음.
  - 읽은 schedule 은 dxf_parser.detect_wall_openings() 에 먹여 평면 벽 끊김과 폭 매칭 →
    창/문 자동 배치, 또는 preview.py 반자동 배치의 드롭다운 카탈로그로 쓰인다.

schedule 레코드(내부 표준) = {mark, subtype('door'|'window'), width, height, sill, count}
Excel 양식 열 = [부호, 종류, 폭(mm), 높이(mm), 창대높이(mm), 수량, 비고]
"""
import os

# Excel 양식 헤더(고정 순서). 사람이 읽는 한글 열 이름.
SCHEDULE_HEADERS = ["부호", "종류", "폭(mm)", "높이(mm)", "창대높이(mm)", "수량", "비고"]
# 종류 드롭다운 허용값(한글) ↔ 내부 subtype
KIND_LABELS = ["창", "문", "문+창"]
_LABEL_TO_SUBTYPE = {"창": "window", "문": "door", "문+창": "door"}


def _mark_to_kind_label(mark, subtype):
    """부호/subtype → 사람이 읽는 종류 라벨(창/문/문+창)."""
    m = (mark or "").upper()
    if "DW" in m or "문+창" in m:           # ADW 등 = 문+창
        return "문+창"
    if subtype == "window" or m.endswith("W") and "D" not in m:
        return "창"
    if subtype == "door":
        return "문"
    return "창" if subtype == "window" else "문"


def _kind_label_to_subtype(label, mark=""):
    """종류 라벨 → subtype. 라벨 없으면 부호/종횡으로 폴백."""
    if label:
        s = _LABEL_TO_SUBTYPE.get(str(label).strip())
        if s:
            return s
    m = (mark or "").upper()
    if m.endswith("W") and "D" not in m:
        return "window"
    return "door"


def _default_sill(subtype):
    return 900.0 if subtype == "window" else 0.0


def export_schedule_xlsx(schedule, path, title="창호일람표"):
    """schedule 레코드 목록 → Excel 양식(.xlsx). schedule 비면 예시 2행 빈 양식.
    반환: 저장 경로."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "창호일람"

    # 제목 행
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(SCHEDULE_HEADERS))
    tcell = ws.cell(row=1, column=1, value=title)
    tcell.font = Font(size=14, bold=True)
    tcell.alignment = Alignment(horizontal="center", vertical="center")

    # 헤더 행(2행)
    hdr_fill = PatternFill("solid", fgColor="2D6CDF")
    hdr_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="B0B7C3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, name in enumerate(SCHEDULE_HEADERS, start=1):
        cell = ws.cell(row=2, column=c, value=name)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # 데이터 행
    rows = schedule if schedule else [
        {"mark": "AW", "subtype": "window", "width": 7700, "height": 1500,
         "sill": 900, "count": 1},
        {"mark": "ADW", "subtype": "door", "width": 2950, "height": 2400,
         "sill": 0, "count": 1},
    ]
    r = 3
    for s in rows:
        sub = s.get("subtype")
        kind = _mark_to_kind_label(s.get("mark"), sub)
        vals = [s.get("mark", ""), kind, s.get("width"), s.get("height"),
                s.get("sill", _default_sill(sub)), s.get("count", 1),
                s.get("remarks", "")]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if c != 7 else "left")
        r += 1

    # 종류 열 드롭다운(창/문/문+창)
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(KIND_LABELS), allow_blank=True)
    dv.add(f"B3:B{max(r, 200)}")
    ws.add_data_validation(dv)

    # 열 너비 + 헤더 고정
    widths = [10, 9, 10, 11, 13, 8, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A3"

    # 안내 메모(데이터 아래)
    note = ("※ 폭/높이/창대높이는 mm. 종류=창/문/문+창. "
            "평면도(DXF)와 함께 입력하면 벽 끊김 폭과 매칭해 창/문 3D 생성에 사용됩니다.")
    ws.cell(row=r + 1, column=1, value=note).font = Font(italic=True, color="7A8290")

    wb.save(path)
    return path


def load_schedule_xlsx(path):
    """Excel 양식(.xlsx) → schedule 레코드 목록.
    헤더 위치를 자동 탐지(부호/폭 열). 빈 행·합계 행은 건너뜀."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # 헤더 행 탐지: '부호'와 '폭'을 포함한 행
    header_row, col = None, {}
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=min(10, ws.max_row)), start=1):
        labels = [(str(c.value).strip() if c.value is not None else "") for c in row]
        if any("부호" in l for l in labels) and any("폭" in l for l in labels):
            header_row = ri
            for ci, l in enumerate(labels):
                col[l] = ci
            break
    if header_row is None:
        return []

    def find(*keys):
        for k, idx in col.items():
            if any(key in k for key in keys):
                return idx
        return None

    i_mark = find("부호")
    i_kind = find("종류")
    i_w = find("폭")
    i_h = find("높이")
    i_sill = find("창대", "sill")
    i_cnt = find("수량", "개수")
    i_rem = find("비고", "비 고")

    def num(row, idx):
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        if v is None or v == "":
            return None
        try:
            return float(str(v).replace(",", ""))
        except (ValueError, TypeError):
            return None

    out = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row is None:
            continue
        mark = (str(row[i_mark]).strip() if i_mark is not None
                and i_mark < len(row) and row[i_mark] is not None else "")
        w = num(row, i_w)
        h = num(row, i_h)
        if w is None or h is None:        # 폭·높이 없으면 유효 행 아님
            continue
        kind = (str(row[i_kind]).strip() if i_kind is not None
                and i_kind < len(row) and row[i_kind] is not None else "")
        sub = _kind_label_to_subtype(kind, mark)
        sill = num(row, i_sill)
        cnt = num(row, i_cnt)
        rem = (str(row[i_rem]).strip() if i_rem is not None
               and i_rem < len(row) and row[i_rem] is not None else "")
        out.append({
            "mark": mark or "?",
            "subtype": sub,
            "width": round(w, 1),
            "height": round(h, 1),
            "sill": float(sill) if sill is not None else _default_sill(sub),
            "count": int(cnt) if cnt is not None else 1,
            **({"remarks": rem} if rem else {}),
        })
    return out


def merge_schedules(base, extra):
    """두 schedule 병합. (mark,width,height) 키 동일하면 extra 우선(사용자 수정본).
    base=DXF 자동추출, extra=Excel 사용자본 → 합집합(extra 가 덮어씀)."""
    by_key = {}
    for s in base or []:
        by_key[(s["mark"], s["width"], s["height"])] = dict(s)
    for s in extra or []:
        by_key[(s["mark"], s["width"], s["height"])] = dict(s)
    return sorted(by_key.values(), key=lambda r: (-r["width"], -r["height"]))


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description="창호일람 Excel 양식 생성/검사")
    ap.add_argument("--blank", metavar="OUT.xlsx", help="빈 양식(예시 2행) 생성")
    ap.add_argument("--from-dxf", metavar="SCHED.dxf",
                    help="창호일람표 DXF에서 추출해 Excel 로 저장(--out 필요)")
    ap.add_argument("--out", metavar="OUT.xlsx", help="출력 .xlsx 경로")
    ap.add_argument("--check", metavar="IN.xlsx", help="Excel 양식 읽어 schedule 출력")
    args = ap.parse_args()

    if args.blank:
        print("빈 양식 생성:", export_schedule_xlsx([], args.blank))
    if args.from_dxf:
        import ezdxf
        import dxf_parser as P
        doc = ezdxf.readfile(args.from_dxf)
        _, sched = P.extract_window_schedule(doc.modelspace(), 1.0)
        out = args.out or (os.path.splitext(args.from_dxf)[0] + "_창호일람.xlsx")
        print(f"추출 {len(sched)}행 → {export_schedule_xlsx(sched, out)}")
    if args.check:
        for s in load_schedule_xlsx(args.check):
            print(s)
